# -*- coding: utf-8 -*-
"""
Final Nystagmus and Eye Tremor Analysis System
Version: 4.0 (Professional, Single-File, OOP, Type-Hinted)

This script is a robust, single-file application for real-time eye tracking and
tremor analysis. It incorporates professional software engineering practices:
 - Object-Oriented structure (Application, EyeTracker classes)
 - Type Hinting for clarity and static analysis
 - Dataclasses for robust data structures
 - try...finally for guaranteed resource cleanup and data saving
 - Encapsulated Logic, Centralized Configuration, Refactored Kinematics (DRY).
 - Non-Blocking Model Training via keypress.
 - Advanced Feature Engineering & Robust Data Quality.
 - Professional Logging & Comprehensive Model Evaluation.
"""

import time
import cv2
import numpy as np
import numpy.typing as npt
import datetime
import pandas as pd
import os
import collections
import logging
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional, Deque, Any, Union
# Add this with the other imports at the top of the file (around line 20)
import tkinter as tk
from tkinter import filedialog

# --- Suppress TensorFlow warnings (must be done before importing TF-dependent modules) ---
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

# Attempt imports, allow for graceful failure if libs missing (better for analysis)
try:
    import mediapipe as mp
    from openpyxl import Workbook, load_workbook, worksheet
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.model_selection import train_test_split
    from sklearn.preprocessing import StandardScaler
    from sklearn.metrics import classification_report
    from joblib import dump, load
    # Define types for ML objects
    SklearnModel = Any 
    SklearnScaler = Any
    MediapipeFaceMesh = Any
except ImportError as e:
     logging.error(f"Missing dependency: {e}. Please install: pip install opencv-python numpy pandas mediapipe openpyxl scikit-learn joblib")
     # Define placeholder types if imports fail
     SklearnModel = Any 
     SklearnScaler = Any
     MediapipeFaceMesh = Any
     Workbook = Any
     worksheet = Any # type: ignore
     exit(1)


# --- Basic Logging Configuration ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


# =============================================================================
# --- 1. CENTRALIZED CONFIGURATION ---
# =============================================================================
# All tunable parameters are here for easy modification.

# --- I/O and File Paths ---
OUTPUT_DIR: str = r"F:\GP\ML\LiveData" # USE YOUR PATH
NEW_DATA_FILENAME: str = "LiveData.xlsx"
OLD_DATA_FILENAME: str = "Live-Old-Data.xlsx"
MODEL_FILENAME: str = "tremor_model.joblib"
SCALER_FILENAME: str = "tremor_scaler.joblib"

# --- Camera and Processing Settings ---
USE_WEBCAM: bool = True
WEBCAM_INDEX: int = 0
VIDEO_FILE_PATH: str = "path/to/your/video.mp4"

# --- MediaPipe Landmark Indices ---
# Using named constants for clarity. "Left" and "Right" are from the person's perspective.
LEFT_IRIS_CENTER: int = 473  # Subject's Left Eye Iris
RIGHT_IRIS_CENTER: int = 468 # Subject's Right Eye Iris
# Used for scaling factor - points on outer edges of eyes
LEFT_EYE_SCALING_PT: int = 362 # Outer corner Subject's Left Eye
RIGHT_EYE_SCALING_PT: int = 263 # Outer corner Subject's Right Eye


# --- Data Quality and Feature Engineering ---
FRAME_WARMUP_PERIOD: int = 15 # Frames to let camera/tracking stabilize
AVG_EYE_WIDTH_MM: float = 24.0 # Used for scaling normalized coordinates to mm
HISTORY_LEN: int = 30          # Frame window for Variance features
VELOCITY_SMOOTHING_WINDOW: int = 5 # Frame window for Simple Moving Average (velocity, acceleration)
VELOCITY_ZERO_THRESHOLD: float = 0.001 # Threshold for zero-crossing frequency count

# --- Machine Learning Feature List ---
# NOTE: The names here MUST match the keys generated in EyeTracker.get_feature_dict
ML_FEATURES_LIST: List[str] = [
    'Left_Move_Magnitude_mm', 'Right_Move_Magnitude_mm', 'Left_Vel_Magnitude_mm_s', 'Right_Vel_Magnitude_mm_s',
    'Left_Freq_Hz', 'Right_Freq_Hz', 'Left_DX_Norm', 'Left_DY_Norm', 'Left_DZ_Norm',
    'Left_VX_Norm_Smoothed', 'Left_VY_Norm_Smoothed', 'Left_VZ_Norm_Smoothed',
    'Right_DX_Norm', 'Right_DY_Norm', 'Right_DZ_Norm',
    'Right_VX_Norm_Smoothed', 'Right_VY_Norm_Smoothed', 'Right_VZ_Norm_Smoothed',
    'Left_DX_Variance', 'Left_DY_Variance', 'Left_DZ_Variance', 'Right_DX_Variance', 'Right_DY_Variance', 'Right_DZ_Variance',
    'Left_AX_Norm_Smoothed', 'Left_AY_Norm_Smoothed', 'Left_AZ_Norm_Smoothed',
    'Right_AX_Norm_Smoothed', 'Right_AY_Norm_Smoothed', 'Right_AZ_Norm_Smoothed'
]
ML_CLASS_WEIGHT: str = 'balanced' # Handle imbalanced datasets
ML_RANDOM_STATE: int = 42

# --- Visualization ---
PLOT_WIDTH: int = 600
PLOT_HEIGHT: int = 300
PLOT_HISTORY_LEN: int = 150 # How many points to show on the real-time graph
MIN_V_PLOT: float = -0.05 # Y-axis range for velocity plot
MAX_V_PLOT: float = 0.05  # Y-axis range for velocity plot

# Type Aliases
Point3D = Tuple[float, float, float]

# =============================================================================
# --- 2. DATA STRUCTURES (Dataclasses) ---
# =============================================================================
@dataclass
class KinematicState:
    """Robust data structure for a single frame's kinematic results."""
    # --- Raw / Base ---
    dx_norm: float = 0.0
    dy_norm: float = 0.0
    dz_norm: float = 0.0
    vx_norm_raw: float = 0.0
     # --- Smoothed ---
    vx_norm_smoothed: float = 0.0
    vy_norm_smoothed: float = 0.0
    vz_norm_smoothed: float = 0.0
    ax_norm_smoothed: float = 0.0
    ay_norm_smoothed: float = 0.0
    az_norm_smoothed: float = 0.0
    # --- Derived / Scaled ---
    move_magnitude_mm: float = 0.0
    vel_magnitude_mm_s: float = 0.0
    freq_hz: float = 0.0
    # --- Variance ---
    dx_variance: float = 0.0
    dy_variance: float = 0.0
    dz_variance: float = 0.0


# =============================================================================
# --- 3. HELPER CLASSES & FUNCTIONS ---
# =============================================================================

class EyeTracker:
    """
    Encapsulates state and kinematic processing for a single eye.
    Manages history, previous state, and calculations (DRY).
     """
    def __init__(self,
                 eye_label: str,
                 history_len: int,
                 smoothing_window: int,
                 zero_threshold: float):
        self.label = eye_label
        self.zero_threshold = zero_threshold
        # State variables
        self.prev_pos: Point3D = (0.0, 0.0, 0.0)
        self.prev_vel_smoothed: Point3D = (0.0, 0.0, 0.0)
        self.prev_vx_for_freq: float = 0.0
        self.oscillation_count: int = 0
         # History Deques for variance
        self.dx_hist: Deque[float] = collections.deque(maxlen=history_len)
        self.dy_hist: Deque[float] = collections.deque(maxlen=history_len)
        self.dz_hist: Deque[float] = collections.deque(maxlen=history_len)
         # History Deques for smoothing
        self.vx_hist: Deque[float] = collections.deque(maxlen=smoothing_window)
        self.vy_hist: Deque[float] = collections.deque(maxlen=smoothing_window)
        self.vz_hist: Deque[float] = collections.deque(maxlen=smoothing_window)

    def process_frame(self,
                      current_pos: Point3D,
                      dt: float,
                      scaling_factor: float,
                      time_with_face: float) -> KinematicState:
        """
        Calculates displacement, velocity, acceleration, variance, frequency.
        Updates internal state and returns results.
        """
        state = KinematicState()
        if dt <= 1e-6: # Avoid division by zero or near-zero dt
             # Still update position to avoid jump on next valid frame
            self.prev_pos = current_pos
            return state # Return default zero state

        # 1. Displacement
        dx, dy, dz = np.array(current_pos) - np.array(self.prev_pos)
        state.dx_norm, state.dy_norm, state.dz_norm = dx, dy, dz
        self.dx_hist.append(dx)
        self.dy_hist.append(dy)
        self.dz_hist.append(dz)

        # 2. Raw and Smoothed Velocity
        vx, vy, vz = dx / dt, dy / dt, dz / dt
        state.vx_norm_raw = vx
        self.vx_hist.append(vx)
        self.vy_hist.append(vy)
        self.vz_hist.append(vz)
        # Simple Moving Average (SMA)
        vx_s = float(np.mean(self.vx_hist)) if self.vx_hist else 0.0
        vy_s = float(np.mean(self.vy_hist)) if self.vy_hist else 0.0
        vz_s = float(np.mean(self.vz_hist)) if self.vz_hist else 0.0
        state.vx_norm_smoothed, state.vy_norm_smoothed, state.vz_norm_smoothed = vx_s, vy_s, vz_s

        # 3. Smoothed Acceleration (based on smoothed velocity)
        state.ax_norm_smoothed = (vx_s - self.prev_vel_smoothed[0]) / dt
        state.ay_norm_smoothed = (vy_s - self.prev_vel_smoothed[1]) / dt
        state.az_norm_smoothed = (vz_s - self.prev_vel_smoothed[2]) / dt

        # 4. Magnitudes (mm and mm/s)
        state.move_magnitude_mm = float(np.linalg.norm([dx, dy, dz])) * scaling_factor
        state.vel_magnitude_mm_s = float(np.linalg.norm([vx_s, vy_s, vz_s])) * scaling_factor

        # 5. Frequency (Zero-Crossing detection on RAW velocity X)
        cross_up = (vx > self.zero_threshold and self.prev_vx_for_freq < -self.zero_threshold)
        cross_down = (vx < -self.zero_threshold and self.prev_vx_for_freq > self.zero_threshold)
        if cross_up or cross_down:
             self.oscillation_count += 1
        # A full cycle is 2 crossings; Hz = cycles / second
        state.freq_hz = (self.oscillation_count / 2) / time_with_face if time_with_face > 0.5 else 0.0 # Wait 0.5s

        # 6. Variance
        state.dx_variance = float(np.var(self.dx_hist)) if self.dx_hist else 0.0
        state.dy_variance = float(np.var(self.dy_hist)) if self.dy_hist else 0.0
        state.dz_variance = float(np.var(self.dz_hist)) if self.dz_hist else 0.0
        
        # 7. IMPORTANT: Update internal state for the *next* frame
        self.prev_pos = current_pos
        self.prev_vel_smoothed = (vx_s, vy_s, vz_s)
        self.prev_vx_for_freq = vx # use raw vx for next zero-crossing check

        return state

    def get_feature_dict(self, state: KinematicState) -> Dict[str, float]:
         """Converts KinematicState + Label to dictionary for ML feature list."""
         # Use attribute access (state.dx_norm) which is safer than dict.get()
         return {
            f'{self.label}_Move_Magnitude_mm': state.move_magnitude_mm,
            f'{self.label}_Vel_Magnitude_mm_s': state.vel_magnitude_mm_s,
            f'{self.label}_Freq_Hz': state.freq_hz,
            f'{self.label}_DX_Norm': state.dx_norm,
            f'{self.label}_DY_Norm': state.dy_norm,
            f'{self.label}_DZ_Norm': state.dz_norm,
            f'{self.label}_VX_Norm_Smoothed': state.vx_norm_smoothed,
            f'{self.label}_VY_Norm_Smoothed': state.vy_norm_smoothed,
            f'{self.label}_VZ_Norm_Smoothed': state.vz_norm_smoothed,
            f'{self.label}_DX_Variance': state.dx_variance,
            f'{self.label}_DY_Variance': state.dy_variance,
            f'{self.label}_DZ_Variance': state.dz_variance,
             f'{self.label}_AX_Norm_Smoothed': state.ax_norm_smoothed,
             f'{self.label}_AY_Norm_Smoothed': state.ay_norm_smoothed,
             f'{self.label}_AZ_Norm_Smoothed': state.az_norm_smoothed,
         }

# --- ML Functions ---

def load_model_and_scaler(
      model_path: str,
      scaler_path: str,
      feature_count: int
      ) -> Tuple[Optional[SklearnModel], Optional[SklearnScaler]]:
      """Loads model and scaler, validates feature count."""
      model, scaler = None, None
      try:
          model_load, scaler_load = load(model_path), load(scaler_path)
          # Validation
          if not hasattr(scaler_load, 'n_features_in_') or scaler_load.n_features_in_ != feature_count:
              raise ValueError(f"Scaler feature count ({getattr(scaler_load, 'n_features_in_', 'N/A')}) mismatch. Expected {feature_count}. Retraining required.")
          if not hasattr(model_load, 'n_features_in_') or model_load.n_features_in_ != feature_count:
               raise ValueError(f"Model feature count ({getattr(model_load, 'n_features_in_', 'N/A')}) mismatch. Expected {feature_count}. Retraining required.")

          model, scaler = model_load, scaler_load
          logging.info(f"Loaded pre-trained model and scaler successfully from {model_path}.")
      except FileNotFoundError:
           logging.warning(f"Model or scaler file not found at {os.path.dirname(model_path)}.")
      except Exception as e:
          logging.warning(f"Could not load or validate pre-trained model/scaler ({e}).")
      return model, scaler


def train_new_model(
        data_paths: List[str],
        feature_list: List[str],
        model_path: str,
        scaler_path: str
       ) -> Tuple[Optional[SklearnModel], Optional[SklearnScaler]]:
    """Loads data, cleans it, and trains/saves a new ML model & scaler."""
    logging.info("--- Attempting to Train New Model ---")
    combined_df = pd.DataFrame()
    valid_sheets_found = 0
    for path in data_paths:
        if not os.path.exists(path):
            logging.warning(f"Training data file not found: {path}. Skipping.")
            continue
        try:
            # Check if file is a valid Excel file
            excel_file = pd.ExcelFile(path)
            sheet_names = [s for s in excel_file.sheet_names if s.startswith("Session_")]
            if not sheet_names:
                 logging.warning(f"No 'Session_' sheets found in {path}. Skipping.")
                 continue
                 
            for sheet_name in sheet_names:
                # IMPORTANT: Skip first calculated row which can have initialization artifacts.
                # Row 0=header, Row 1=first data, which might be a jump from (0,0,0)
                sheet_df = pd.read_excel(path, sheet_name=sheet_name, skiprows=[1]) 
                if not sheet_df.empty and 'Label' in sheet_df.columns:
                    combined_df = pd.concat([combined_df, sheet_df], ignore_index=True)
                    valid_sheets_found +=1
        except Exception as e:
            logging.error(f"Could not read or parse Excel file {path}: {e}. Skipping.")

    if combined_df.empty or valid_sheets_found == 0:
        logging.error("No valid training data found across all paths/sheets. Model training aborted.")
        return None, None

    # Data Cleaning & Validation
    missing_features = [f for f in feature_list if f not in combined_df.columns]
    for feature in missing_features:
        logging.warning(f"Feature '{feature}' not in data. Filling with 0.0.")
        combined_df[feature] = 0.0
        
    if 'Label' not in combined_df.columns:
        logging.error("'Label' column not found in data. Cannot train. Aborting.")
        return None, None
        
    # Drop rows with ANY NaN in the features or label
    original_rows = len(combined_df)
    combined_df.dropna(subset=feature_list + ['Label'], inplace=True)
    # Replace infinite values with NaN and drop again
    combined_df.replace([np.inf, -np.inf], np.nan, inplace=True)
    combined_df.dropna(subset=feature_list + ['Label'], inplace=True)
    
    logging.info(f"Training rows: {original_rows} original, {len(combined_df)} after cleaning (NaN/Inf removal).")

    num_classes = combined_df['Label'].nunique()
    if len(combined_df) < 50 or num_classes < 2:
        logging.error(f"Insufficient data ({len(combined_df)} rows) or classes ({num_classes}) for robust training. Need >= 50 rows and >= 2 classes.")
        return None, None

    X = combined_df[feature_list]
    y = combined_df['Label'].astype(int) # Ensure labels are integers
    logging.info(f"Training data label distribution:\n{y.value_counts().to_string()}")

    try:
       # stratify ensures train/test sets have similar class proportions
       X_train, X_test, y_train, y_test = train_test_split(
           X, y, test_size=0.25, random_state=ML_RANDOM_STATE, stratify=y, shuffle=True)
    except ValueError as e:
        logging.error(f"Stratified split failed (e.g., a class has only 1 member): {e}. Aborting.")
        return None, None
        
    scaler = StandardScaler().fit(X_train)
    model = RandomForestClassifier(
        n_estimators=100, max_depth=12, random_state=ML_RANDOM_STATE, class_weight=ML_CLASS_WEIGHT, n_jobs=-1)
    
    X_train_scaled = scaler.transform(X_train)
    X_test_scaled = scaler.transform(X_test)
    
    model.fit(X_train_scaled, y_train)

    # Professional Evaluation
    y_pred = model.predict(X_test_scaled)
    # zero_division=0 prevents warnings if a class is not predicted at all
    report = classification_report(y_test, y_pred, zero_division=0) 
    logging.info(f"\n--- Model Evaluation Report (on Test Set) ---\n{report}")

    # Save artifacts
    try:
       os.makedirs(os.path.dirname(model_path), exist_ok=True)
       dump(model, model_path)
       dump(scaler, scaler_path)
       logging.info(f"New model and scaler saved to {os.path.dirname(model_path)}")
    except Exception as e:
       logging.error(f"Failed to save model/scaler to {os.path.dirname(model_path)}: {e}")
       # return model anyway so app can use it for the current session
       
    return model, scaler

# --- Data Saving Function ---
def save_data_to_excel(output_path: str, data_rows: List[List[Any]]) -> None:
     """Saves collected data frame-by-frame to a new sheet in an Excel file."""
     if not data_rows or len(data_rows) <=1: # Only header or no data
         logging.warning("No data collected to save.")
         return
     try:
        logging.info(f"Attempting to save {len(data_rows)-1} data rows to {output_path}...")
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        wb: Workbook
        if os.path.exists(output_path):
            wb = load_workbook(output_path)
        else:
            wb = Workbook()
            # Remove default sheet if it exists
            if "Sheet" in wb.sheetnames:
                 wb.remove(wb["Sheet"]) 

        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        # Ensure sheet name is unique if saving very quickly
        sheet_title_base = f"Session_{timestamp}"
        sheet_title = sheet_title_base
        counter = 1
        while sheet_title in wb.sheetnames:
            sheet_title = f"{sheet_title_base}_{counter}"
            counter +=1
            
        ws: worksheet.worksheet.Worksheet = wb.create_sheet(title=sheet_title)
        for row in data_rows:
            ws.append(row)
        wb.save(output_path)
        logging.info(f"Data saved successfully to sheet '{sheet_title}'.")
     except PermissionError:
          logging.critical(f"FATAL: Permission denied saving {output_path}. Is the file open elsewhere?")
     except Exception as e:
        logging.critical(f"FATAL: Could not save to Excel file {output_path}. Error: {e}", exc_info=True)

# --- Visualization Function ---
def normalize_for_plot(value: float, min_val: float, max_val: float, height: int) -> int:
    """Scales a data value to fit within the pixel height of the plot."""
    # Clip value to range to prevent lines going off canvas
    value_clipped = max(min_val, min(max_val, value))
    range_val = max_val - min_val
    if range_val < 1e-6: # Avoid division by zero if min==max
        return height // 2
    # Calculate position and invert (0 is top in OpenCV)
    return int(height - ((value_clipped - min_val) / range_val) * height)


# =============================================================================
# --- 4. MAIN APPLICATION CLASS ---
# =============================================================================

class Application:
    """Manages the main application loop, resources, and state."""
    def __init__(self):
        logging.info("Initializing Nystagmus Analysis Application...")
        self.output_dir = OUTPUT_DIR
        self.data_paths: List[str] = [os.path.join(self.output_dir, NEW_DATA_FILENAME),
                                      os.path.join(self.output_dir, OLD_DATA_FILENAME)]
        self.model_path: str = os.path.join(self.output_dir, MODEL_FILENAME)
        self.scaler_path: str = os.path.join(self.output_dir, SCALER_FILENAME)
        self.save_path: str = os.path.join(self.output_dir, NEW_DATA_FILENAME)

        self.mp_face_mesh = mp.solutions.face_mesh
        self.face_mesh: Optional[MediapipeFaceMesh] = None
        self.cap: Optional[cv2.VideoCapture] = None
        self.model: Optional[SklearnModel] = None
        self.scaler: Optional[SklearnScaler] = None

        # Instantiate Trackers
        self.tracker_left = EyeTracker("Left", HISTORY_LEN, VELOCITY_SMOOTHING_WINDOW, VELOCITY_ZERO_THRESHOLD)
        self.tracker_right = EyeTracker("Right", HISTORY_LEN, VELOCITY_SMOOTHING_WINDOW, VELOCITY_ZERO_THRESHOLD)
        
        # State
        self.frame_index: int = 0
        self.total_time: float = 0.0
        self.time_with_face: float = 0.0
        self.last_timestamp: float = time.time()
        
        # Data Collection & Plotting
        self.excel_header: List[str] = ["Index", "TotalTime", "TimeWithFace"] + ML_FEATURES_LIST + ["Prediction", "Confidence", "Label"]
        self.data_for_excel: List[List[Any]] = [self.excel_header]
        self.plot_history: Dict[str, Deque[float]] = {
           # Store smoothed values for plotting
           'vx_l': collections.deque(maxlen=PLOT_HISTORY_LEN), 'vy_l': collections.deque(maxlen=PLOT_HISTORY_LEN),
           'vx_r': collections.deque(maxlen=PLOT_HISTORY_LEN), 'vy_r': collections.deque(maxlen=PLOT_HISTORY_LEN)
         }
        self.plot_colors: List[Tuple[int,int,int]] = [(0,255,0), (255,0,0), (0,255,255), (255,0,255)] # L-VX, L-VY, R-VX, R-VY (BGR)
        self.window_main: str = "Nystagmus Detection"
        self.window_plot: str = "Smoothed Eye Velocity (X=Green/Yellow, Y=Red/Magenta)"


    def _initialize_resources(self) -> bool:
        """Setup camera and mediapipe."""
        try:
             self.face_mesh = self.mp_face_mesh.FaceMesh(
                 static_image_mode=False, max_num_faces=1, refine_landmarks=True,
                 min_detection_confidence=0.005, min_tracking_confidence=0.6)
        except Exception as e:
             logging.critical(f"Failed to initialize MediaPipe: {e}")
             return False

        # ----------------------------------------------- 
        # OPTION 1 (commented out by default): Use webcam 
        self.cap = cv2.VideoCapture(WEBCAM_INDEX)
        if not self.cap.isOpened():
            logging.critical(f"Error: Could not open webcam index {WEBCAM_INDEX}.")
            return False
        source = f"webcam index {WEBCAM_INDEX}"
        
        # OPTION 2: Use a file dialog to select any video
        # root = tk.Tk()
        # root.withdraw()  # Hide the main tkinter window
        
        # video_path = filedialog.askopenfilename(
        #     title="Select a video file",
        #     filetypes=[("Video Files", "*.mp4 *.avi *.mov *.mkv"), ("All Files", "*.*")]
        # )
        
        # if not video_path:
        #     logging.critical("No file selected. Exiting...")
        #     return False
        
        # self.cap = cv2.VideoCapture(video_path)
        # source = video_path
        # -----------------------------------------------
            
        if not self.cap or not self.cap.isOpened():
            logging.critical(f"Cannot open video source: {source}.")
            self.cap = None # Ensure it's None if failed
            return False
        logging.info(f"Video source '{source}' opened successfully.")
        return True

    def _release_resources(self) -> None:
       """Clean up camera, windows, mediapipe."""
       logging.info("Releasing resources...")
       if self.cap:
           self.cap.release()
       if self.face_mesh:
            self.face_mesh.close() # type: ignore
       cv2.destroyAllWindows()

    def _handle_training(self, frame: npt.NDArray) -> None:
        """Triggered by keypress to train model."""
        logging.info("--- Training key pressed. Attempting to train a new model... ---")
        h, w, _ = frame.shape
        cv2.putText(frame, "TRAINING MODEL...", (w//4 , h//2), cv2.FONT_HERSHEY_SIMPLEX, 1.5, (0, 0, 255), 3)
        cv2.imshow(self.window_main, frame)
        cv2.waitKey(1) # Force window update to show text
        
        new_model, new_scaler = train_new_model(
             self.data_paths, ML_FEATURES_LIST, self.model_path, self.scaler_path)
             
        if new_model and new_scaler:
            self.model, self.scaler = new_model, new_scaler
            logging.info("Training successful. New model is now active.")
        else:
            logging.error("Training failed. Continuing with previous model (if any).")

    def _update_plot(self, l_state: KinematicState, r_state: KinematicState) -> npt.NDArray:
         """Draw the real-time velocity plot."""
         plot_img = np.zeros((PLOT_HEIGHT, PLOT_WIDTH, 3), dtype=np.uint8)
         # Draw zero line
         cv2.line(plot_img, (0, PLOT_HEIGHT//2), (PLOT_WIDTH, PLOT_HEIGHT//2), (70,70,70), 1)
         
         self.plot_history['vx_l'].append(l_state.vx_norm_smoothed)
         self.plot_history['vy_l'].append(l_state.vy_norm_smoothed)
         self.plot_history['vx_r'].append(r_state.vx_norm_smoothed)
         self.plot_history['vy_r'].append(r_state.vy_norm_smoothed)
         
         x_scale = PLOT_WIDTH / PLOT_HISTORY_LEN
         for i, (key, hist) in enumerate(self.plot_history.items()):
            if len(hist) < 2: continue
            points = []
            for j, v in enumerate(hist):
                 x = int(j * x_scale)
                 y = normalize_for_plot(v, MIN_V_PLOT, MAX_V_PLOT, PLOT_HEIGHT)
                 points.append([x,y])
            
            points_arr = np.array(points, dtype=np.int32)
            cv2.polylines(plot_img, [points_arr], isClosed=False, color=self.plot_colors[i], thickness=1)
         return plot_img

    def run(self) -> None:
        """Main application execution loop with try...finally for cleanup."""
        if not self._initialize_resources():
            self._release_resources()
            return
            
        self.model, self.scaler = load_model_and_scaler(self.model_path, self.scaler_path, len(ML_FEATURES_LIST))
        if not self.model or not self.scaler:
             logging.warning("Model not ready. Run in data-collection mode or press 'T' to train.")

        self.last_timestamp = time.time()
        l_state, r_state = KinematicState(), KinematicState() # Default states
        
        try: # --- TRY BLOCK: Main Loop ---
            while self.cap and self.cap.isOpened():
                ret, frame = self.cap.read()
                if not ret:
                    logging.info("End of video stream.")
                    break

                frame = cv2.flip(frame, 1) # Flip horizontally for mirror view
                h, w, _ = frame.shape
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                
                # Time calculation
                current_timestamp = time.time()
                dt = current_timestamp - self.last_timestamp
                # Prevent excessive dt if paused/lagged, cap at 0.1s (10 fps)
                dt = min(dt, 0.1) 
                self.last_timestamp = current_timestamp
                self.total_time += dt
                fps = 1.0 / dt if dt > 0 else 0.0
                
                # Reset per-frame state
                prediction_val: Union[str, int, float] = "No Face"
                prediction_conf: float = 0.0
                all_features: Dict[str, float] = {feature: 0.0 for feature in ML_FEATURES_LIST}
                face_detected = False
                is_warmup = self.frame_index < FRAME_WARMUP_PERIOD

                # Mediapipe Processing
                results = self.face_mesh.process(frame_rgb) # type: ignore

                if results and results.multi_face_landmarks:
                     face_detected = True
                     landmarks = results.multi_face_landmarks[0].landmark
                     
                     # Get points
                     l_center: Point3D = (landmarks[LEFT_IRIS_CENTER].x, landmarks[LEFT_IRIS_CENTER].y, landmarks[LEFT_IRIS_CENTER].z)
                     r_center: Point3D = (landmarks[RIGHT_IRIS_CENTER].x, landmarks[RIGHT_IRIS_CENTER].y, landmarks[RIGHT_IRIS_CENTER].z)
                     p_left_scale = landmarks[LEFT_EYE_SCALING_PT]
                     p_right_scale = landmarks[RIGHT_EYE_SCALING_PT]
                    
                     # Calculate Scaling Factor (using 3D distance between eye points)
                     eye_dist_norm = np.linalg.norm(
                         np.array([p_left_scale.x, p_left_scale.y, p_left_scale.z]) - 
                         np.array([p_right_scale.x, p_right_scale.y, p_right_scale.z])
                         )
                     scaling_factor = AVG_EYE_WIDTH_MM / eye_dist_norm if eye_dist_norm > 1e-6 else 0.0
                     
                     # Only increment time_with_face after warmup
                     current_time_with_face = self.time_with_face if is_warmup else self.time_with_face + dt
                     if not is_warmup:
                          self.time_with_face = current_time_with_face
                          prediction_val = "Tracking" # Default if tracking but no model

                     # Process Eyes - MUST run even during warmup to build history and update prev_pos
                     l_state = self.tracker_left.process_frame(l_center, dt, scaling_factor, current_time_with_face)
                     r_state = self.tracker_right.process_frame(r_center, dt, scaling_factor, current_time_with_face)

                     if is_warmup:
                          prediction_val = "Calibrating..."
                     else:
                         # --- Assemble Feature Dictionary ---
                         all_features.update(self.tracker_left.get_feature_dict(l_state))
                         all_features.update(self.tracker_right.get_feature_dict(r_state))

                         # --- Prediction ---
                         if self.model and self.scaler:
                            try:
                                feature_vector_list = [all_features.get(f, 0.0) for f in ML_FEATURES_LIST]
                                feature_vector: npt.NDArray = np.array(feature_vector_list).reshape(1, -1) # Reshape for single sample
                                scaled_vector = self.scaler.transform(feature_vector)
                                prediction_val = self.model.predict(scaled_vector)[0]
                                prediction_conf = float(np.max(self.model.predict_proba(scaled_vector)))
                            except Exception as e:
                                 logging.warning(f"Prediction failed on frame {self.frame_index}: {e}")
                                 prediction_val = "Pred_Error"
                         else:
                              prediction_val = "No Model"

                # --- Data Logging (even if no face, log the zeros) ---
                # Label defaults to 0, adjust if you have a keypress to set labels during recording
                excel_row_dict = {"Index": self.frame_index, "TotalTime": self.total_time, 
                                  "TimeWithFace": self.time_with_face, "Prediction": prediction_val,
                                   "Confidence": prediction_conf, "Label": 0} 
                excel_row_dict.update(all_features)
                # Ensure row order matches header exactly
                self.data_for_excel.append([excel_row_dict.get(h, 0.0) for h in self.excel_header])

                # --- Visualization ---
                plot_img = self._update_plot(l_state, r_state)
                cv2.imshow(self.window_plot, plot_img)

                # Display Overlay on main frame
                display_prefix = "" if face_detected or is_warmup else "[NO FACE] "
                display_text = f"{display_prefix}State: {prediction_val}"
                if isinstance(prediction_val, (int,float)) or (prediction_val not in ["No Face", "Calibrating...", "No Model", "Pred_Error", "Tracking"]):
                   display_text += f" (Conf: {prediction_conf:.2f})"
                   
                color = (0, 165, 255) if is_warmup else (0, 255, 0) # Orange for warmup, Green otherwise
                if not face_detected: color = (0,0,200) # Red if no face
                cv2.putText(frame, display_text, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)
                cv2.putText(frame, f"FPS: {fps:.1f}", (w - 120, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (200, 200, 200), 2)
                cv2.putText(frame, f"Frame: {self.frame_index}", (w - 120, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (200, 200, 200), 1)
                if not self.model or not self.scaler:
                     cv2.putText(frame, "Press 'T' to Train Model", (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2) # Yellow
                cv2.imshow(self.window_main, frame)
                
                self.frame_index += 1

                # --- Key Handling ---
                key = cv2.waitKey(1) & 0xFF
                if key == 27 or key == ord('q'): # ESC or Q Key
                    logging.info("Exit key pressed.")
                    break
                elif key == ord('t'):
                     self._handle_training(frame.copy()) # Pass copy in case training takes time

        except KeyboardInterrupt:
             logging.warning("Ctrl+C detected.")
        except Exception as e:
            logging.error(f"An unhandled error occurred in the main loop at frame {self.frame_index}: {e}", exc_info=True)
            
        finally: # --- FINALLY BLOCK: Always Run ---
            logging.info("Shutting down application...")
            self._release_resources()
            
            # Save data ONLY if loop finished (break or error), not during initialization failure
            if self.cap: # check if capture was ever successfully created
                
                # --- [MODIFIED CODE] SAVE FINAL PLOT IMAGE ---
                try:
                    # Check if there's meaningful data to plot (i.e., loop ran for a bit)
                    if self.frame_index > FRAME_WARMUP_PERIOD:
                        logging.info("Generating final signal plot for saving...")
                        
                        # l_state and r_state are available from the last iteration before the loop exited.
                        final_plot_image = self._update_plot(l_state, r_state)

                        # Create a unique filename. This timestamp will be very close to the one used for the Excel sheet.
                        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
                        plot_filename = f"SignalPlot_{timestamp}.png"
                        
                        # Get the directory from the excel save path to ensure they are in the same location
                        output_directory = os.path.dirname(self.save_path)
                        plot_save_path = os.path.join(output_directory, plot_filename)

                        # Save the image using OpenCV
                        success = cv2.imwrite(plot_save_path, final_plot_image)
                        if success:
                            logging.info(f"Signal plot saved successfully to '{plot_save_path}'.")
                        else:
                            logging.error(f"Failed to save signal plot to '{plot_save_path}'. Check path and permissions.")
                    else:
                        logging.warning("Skipping plot saving: session was too short or no data was processed.")
                except NameError:
                     # This can happen if the loop never ran once (e.g. video file not found)
                     logging.warning("Skipping plot saving: main loop did not run, no state to generate plot from.")
                except Exception as e:
                    logging.error(f"An error occurred while saving the final plot image: {e}", exc_info=True)
                # --- END OF MODIFIED CODE ---

                # Now save the Excel data
                save_data_to_excel(self.save_path, self.data_for_excel)
                
            logging.info("Application terminated.")


# =============================================================================
# --- 5. SCRIPT ENTRY POINT ---
# =============================================================================
def main() -> None:
     """Creates and runs the application instance."""
     app = Application()
     app.run()
    
if __name__ == "__main__":
    main()