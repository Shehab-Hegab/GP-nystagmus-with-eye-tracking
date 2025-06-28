"""
Initial Eye Tracking Analysis and Automated Report Generation System
----------------------------------------------------------

This system provides end-to-end processing of eye-tracking data, from raw data loading to 
PDF report generation. The implementation consists of three core components:

1. DATA LOADER (DataLoader class)
- Handles input from Excel/CSV files with multiple sessions
- Performs data validation and cleaning
- Automatically detects outliers using statistical methods (Z-score thresholding)
- Maintains session isolation for individual analysis

2. DATA VISUALIZER (DataVisualizer class)
- Generates publication-quality visualizations:
  * Time-series grids with outlier highlighting
  * Correlation heatmaps with significance thresholds
  * Movement analysis with configurable thresholds
  * Statistical summaries with IQR/outlier metrics
- Maintains consistent styling through VisualizationConfig

3. REPORT GENERATOR (ReportGenerator class)
- Creates comprehensive PDF reports containing:
  * Cover page with session metadata
  * All visualizations in standardized layout
  * Statistical summaries
- Features:
  * Automatic directory creation
  * Robust path validation
  * Keyboard interrupt handling
  * Error recovery mechanisms

Key Technical Features:
- Config-driven analysis through VisualizationConfig dataclass
- Type hints for maintainability
- Exception handling at all levels
- Memory-efficient PDF generation using matplotlib's PdfPages
- Automated outlier detection and flagging
- Customizable thresholds for movement/outlier detection

Usage Flow:
1. Initialize VisualizationConfig with desired parameters
2. Load data using DataLoader
3. Generate visualizations via DataVisualizer
4. Compile PDF reports with ReportGenerator

Example:
    config = VisualizationConfig(
        excel_file_path="data.xlsx",
        session_to_visualize="Session_123",
        movement_threshold_mm=5.0
    )
    df = DataLoader.load_session_data(config)
    ReportGenerator.generate_pdf_report(df, config)

Dependencies:
- pandas, numpy, matplotlib, seaborn
- openpyxl (for Excel handling)
- keyboard (for interrupt handling)

Note: All paths can be absolute or relative, with automatic validation and fallback.
"""

# --- Imports ---
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import seaborn as sns
from datetime import datetime
import os
from math import ceil
from typing import List, Optional
from dataclasses import dataclass
from scipy import stats
from openpyxl import load_workbook
import warnings
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, 
                               Image, Table, TableStyle, PageBreak)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import tempfile

# --- Configuration ---
@dataclass
class VisualizationConfig:
    excel_file_path: str = ''
    session_to_visualize: str = ''
    x_axis_column: str = 'TimeWithFace'
    plot_style: str = 'seaborn-v0_8-whitegrid'
    grid_plot_cols: int = 3
    grid_plot_figsize: tuple = (20, 5)
    custom_plot_figsize: tuple = (15, 7)
    excluded_columns: List[str] = None
    correlation_threshold: float = 0.7
    outlier_threshold: float = 3.0
    movement_threshold_mm: float = 5.0
    
    def __post_init__(self):
        if self.excluded_columns is None:
            self.excluded_columns = ['Index', 'Label', 'Prediction', 'TotalTime', self.x_axis_column]

# --- Data Loader ---
class DataLoader:
    @staticmethod
    def load_session_data(file_path: str, sheet_name: str, config: VisualizationConfig) -> pd.DataFrame:
        """Excel sheet loader"""
        
        print(f"⚙️ Loading session: '{sheet_name}' with outlier detection...")
        
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            x_axis_col = config.x_axis_column
            
            # Data validation
            if df.empty:
                raise ValueError("Empty DataFrame")
                
            if x_axis_col not in df.columns:
                raise ValueError(f"X-axis column '{x_axis_col}' not found")
            
            # Clean data
            df = df.iloc[1:].copy() if len(df) > 1 else df.copy()
            df = df.reset_index(drop=True)
            
            # Convert x-axis
            df[x_axis_col] = pd.to_numeric(df[x_axis_col], errors='coerce')
            if df[x_axis_col].isnull().any():
                nan_count = df[x_axis_col].isnull().sum()
                df = df.dropna(subset=[x_axis_col]).copy()
                
                print(f"⚠️ Removed {nan_count} rows with invalid time values")
            
            # Optimize numeric columns
            numeric_cols = df.select_dtypes(include=np.number).columns
            df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce')
        
            print(f"✅ Successfully loaded '{sheet_name}' with {len(df)} rows")
            
            return DataLoader.add_outlier_flags(df, config) # Add outlier flags
    
        except Exception as e:
                print(f"❌ Failed to load sheet '{sheet_name}': {str(e)}")
                return None
    
    @staticmethod
    def list_available_sessions(file_path: str) -> List[str]:
        """Lists available sessions"""
        
        if not os.path.exists(file_path):
            print(f"❌ Error: File not found at '{file_path}'")
            return []
            
        try:
            wb = load_workbook(file_path, read_only=True)
            session_sheets = [sheet for sheet in wb.sheetnames if sheet.startswith("Session_")]
            wb.close()
            return sorted(session_sheets)
        except Exception as e:
            print(f"❌ Error reading Excel file: {str(e)}")
            return []
    
    @staticmethod
    def add_outlier_flags(df: pd.DataFrame, config: VisualizationConfig) -> pd.DataFrame:
        """Standalone method to add outlier flags"""
        
        movement_cols = [c for c in df.columns if 'Magnitude' in c]
        
        for col in movement_cols:
            z_scores = np.abs(stats.zscore(df[col]))
            df[f'{col}_Outlier'] = z_scores > config.outlier_threshold
        
        print(f"🔍 Found {df[[f'{c}_Outlier' for c in movement_cols]].sum().sum()} potential outliers")
        
        return df

# --- Data Visualizer ---
class DataVisualizer:
    @staticmethod
    # def _print_used_features(df: pd.DataFrame, config: VisualizationConfig, analysis_type: str):
    #     """Helper method to print features being used for visualization"""
    #     numeric_cols = df.select_dtypes(include=np.number).columns.tolist()
    #     features = [col for col in numeric_cols 
    #                if col not in config.excluded_columns 
    #                and not col.endswith('_Outlier')]
        
    #     print(f"\n📊 Features used in {analysis_type} visualization:")
    #     print(f"X-axis column: '{config.x_axis_column}'")
    #     print("Y-axis features:")
    #     for i, feature in enumerate(features, 1):
    #         print(f"  {i}. {feature}")
        
    #     if config.excluded_columns:
    #         print("\n🚫 Excluded columns:")
    #         for col in config.excluded_columns:
    #             print(f"  - {col}")
        
    #     outlier_cols = [col for col in df.columns if col.endswith('_Outlier')]
    #     if outlier_cols:
    #         print("\n🔴 Outlier detection columns:")
    #         for col in outlier_cols:
    #             print(f"  - {col}")
    #     print("-"*50)
    
    @staticmethod
    def _configure_figure_for_pdf(fig: plt.Figure) -> plt.Figure:
        """Configure matplotlib figure settings for optimal PDF output"""
        fig.tight_layout()
        fig.set_facecolor('white')
        for ax in fig.axes:
            ax.tick_params(axis='both', which='major', labelsize=8)
            ax.xaxis.label.set_size(9)
            ax.yaxis.label.set_size(9)
            ax.title.set_size(10)
        return fig
    
    @staticmethod
    def create_grid_plot(df: pd.DataFrame, config: VisualizationConfig) -> plt.Figure:
        if df is None or df.empty:
            print("⚠️ No data available")
            return None

        numeric_cols = df.select_dtypes(include=np.number).columns.tolist()
        features = [col for col in numeric_cols 
                if col not in config.excluded_columns 
                and not col.endswith('_Outlier')]
        
        # # Print features being used
        # DataVisualizer._print_used_features(df, config, "grid plot")
        
        if not features:
            print("⚠️ No features available")
            return None
            
        n_rows = ceil(len(features) / config.grid_plot_cols)
        fig, axes = plt.subplots(
            n_rows, config.grid_plot_cols,
            figsize=(config.grid_plot_figsize[0], 
                    config.grid_plot_figsize[1] * n_rows),
            squeeze=False
        )
        
        # fig.suptitle(f"Session: {config.session_to_visualize}", fontsize=16)
        
        axes = axes.flatten()
        
        for i, feature in enumerate(features):
            ax = axes[i]
            sns.lineplot(
                data=df, 
                x=config.x_axis_column, 
                y=feature,
                ax=ax,
                linewidth=1,
                color='royalblue'
            )
            
            # Outlier markers
            outlier_col = f'{feature}_Outlier'
            if outlier_col in df.columns:
                outliers = df[df[outlier_col]]
                ax.scatter(
                    outliers[config.x_axis_column],
                    outliers[feature],
                    color='red',
                    s=20,
                    label='Outlier'
                )
            
            ax.set_title(feature, fontsize=12, pad=10)
            ax.set_xlabel(config.x_axis_column, fontsize=10)
            ax.grid(True, linestyle='--', alpha=0.5)
            ax.tick_params(axis='x', labelrotation=45)
        
        # Hide unused subplots
        for j in range(i+1, len(axes)):
            axes[j].axis('off')
        
        plt.tight_layout()
        return DataVisualizer._configure_figure_for_pdf(fig)

    @staticmethod
    def plot_selected_features(df: pd.DataFrame, features: List[str], title: str, config: VisualizationConfig) -> plt.Figure:
        """
        Plots selected features with outliers marked as scatter points.
        Returns matplotlib Figure for PDF generation.
        """
        if df is None or df.empty:
            print("⚠️ No data available for plotting")
            return None
        
        # # Print features being used
        # print(f"\n📊 Features selected for custom visualization '{title}':")
        # print(f"X-axis column: '{config.x_axis_column}'")
        # print("Y-axis features:")
        # for i, feature in enumerate(features, 1):
        #     print(f"  {i}. {feature}")
        # print("-"*50)

        fig, ax = plt.subplots(figsize=config.custom_plot_figsize)
        
        # Plot main lines and outliers
        for feature in features:
            if feature in df.columns:
                # Plot the main line
                ax.plot(
                    df[config.x_axis_column], 
                    df[feature], 
                    label=feature,
                    linewidth=1.5,
                    alpha=0.8
                )
                
                # Add outlier scatter points if available
                outlier_col = f'{feature}_Outlier'
                if outlier_col in df.columns:
                    outliers = df[df[outlier_col]]
                    ax.scatter(
                        outliers[config.x_axis_column],
                        outliers[feature],
                        color='red',
                        s=40,
                        edgecolor='black',
                        linewidth=0.5,
                        label=f'{feature} Outlier',
                        zorder=3  # Ensure points appear above lines
                    )
        
        # Formatting
        ax.set_title(title, fontsize=14, pad=12)
        ax.set_xlabel(config.x_axis_column, fontsize=11)
        ax.set_ylabel("Value", fontsize=11)
        ax.grid(True, alpha=0.3, linestyle='--')
        
        # Legend with duplicate handling
        handles, labels = ax.get_legend_handles_labels()
        unique_labels = dict(zip(labels, handles))  # Removes duplicates
        ax.legend(
            unique_labels.values(),
            unique_labels.keys(),
            bbox_to_anchor=(1.05, 1),
            loc='upper left',
            framealpha=0.9
        )
        
        plt.tight_layout()
        return DataVisualizer._configure_figure_for_pdf(fig)

    @staticmethod
    def plot_movement_analysis(df: pd.DataFrame, config: VisualizationConfig) -> plt.Figure:
        """
        Analyzes movement data with threshold visualization and outlier marking.
        Returns matplotlib Figure optimized for PDF generation.
        """
        if df is None or df.empty:
            print("⚠️ No data available for movement analysis")
            return None

        movement_cols = [c for c in df.columns 
                        if 'Magnitude' in c 
                        and not c.endswith('_Outlier')]
        
        # # Print movement columns being analyzed
        # print("\n📊 Movement analysis features:")
        # print(f"Threshold: {config.movement_threshold_mm} mm")
        # print("Movement columns:")
        # for i, col in enumerate(movement_cols, 1):
        #     print(f"  {i}. {col}")
        # print("-"*50)
        
        if not movement_cols:
            print("⚠️ No movement data columns found")
            return None

        fig, ax = plt.subplots(figsize=(15, 8))
        
        # Custom color cycle for better PDF visibility
        colors = plt.cm.tab10(np.linspace(0, 1, len(movement_cols)))
        
        for idx, col in enumerate(movement_cols):
            # Main movement line
            ax.plot(df[config.x_axis_column], df[col], 
                    label=col, 
                    color=colors[idx],
                    linewidth=1.5,
                    alpha=0.8)
            
            # Threshold line (only label once)
            ax.axhline(y=config.movement_threshold_mm, 
                    color='darkred', 
                    linestyle='--', 
                    linewidth=1.2,
                    label='Threshold' if idx == 0 else "")
            
            # Extreme movements (both outliers and threshold crossings)
            extremes = df[(df[col] > config.movement_threshold_mm) | 
                        (df.get(f'{col}_Outlier', False))]
            if not extremes.empty:
                ax.scatter(extremes[config.x_axis_column], extremes[col],
                        color='red',
                        s=45,
                        edgecolor='black',
                        linewidth=0.8,
                        zorder=3,
                        label='Extreme/Outlier' if idx == 0 else "")

        # Formatting
        ax.set_title(f"Movement Analysis (Threshold = {config.movement_threshold_mm}mm)", 
                    fontsize=14, pad=12)
        ax.set_xlabel(config.x_axis_column, fontsize=11)
        ax.set_ylabel("Movement (mm)", fontsize=11)
        ax.grid(True, alpha=0.3, linestyle=':')
        
        # Smart legend with duplicate handling
        handles, labels = ax.get_legend_handles_labels()
        by_label = dict(zip(labels, handles))
        ax.legend(by_label.values(), by_label.keys(),
                bbox_to_anchor=(1.05, 1),
                loc='upper left',
                framealpha=0.9)

        plt.tight_layout()
        return DataVisualizer._configure_figure_for_pdf(fig)

    @staticmethod
    def create_correlation_heatmap(df: pd.DataFrame, config: VisualizationConfig) -> plt.Figure:
        """
        Generates a correlation heatmap optimized for PDF output.
        Returns matplotlib Figure object.
        """
        # Input validation
        if df is None or df.empty:
            print("⚠️ No data available for heatmap")
            return None

        try:
            # Column selection with more robust filtering
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            exclude_cols = [c for c in numeric_cols 
               if ('_Outlier' in c) 
               or (c == config.x_axis_column) 
               or (c == 'Index') 
               or ('Variance' in c)]
            corr_cols = [c for c in numeric_cols if c not in exclude_cols]
            
            # # Print correlation matrix features
            # print("\n📊 Features included in correlation matrix:")
            # for i, col in enumerate(corr_cols, 1):
            #     print(f"  {i}. {col}")
            # print("-"*50)
            
            if len(corr_cols) < 2:
                print("⚠️ Insufficient numeric columns (need ≥2)")
                return None

            # Calculate correlations with error handling
            corr_matrix = df[corr_cols].corr(method='pearson', numeric_only=True)
            corr_matrix = corr_matrix.round(2)  # Standardize decimal places

            # Create figure with optimized settings
            fig, ax = plt.subplots(figsize=(16, 14))  # Slightly reduced for PDF margins
            
            # Enhanced heatmap visualization
            sns.heatmap(
                corr_matrix,
                ax=ax,
                cmap='coolwarm',  # Better for grayscale printing
                center=0,
                annot=True,  # Now showing values
                annot_kws={'size': 8},
                fmt='.2f',
                linewidths=0.5,
                linecolor='lightgray',
                cbar_kws={'shrink': 0.8},
                square=True  # Equal aspect ratio
            )

            # Improved labeling
            ax.set_title(
                f'Feature Correlations\n(Session: {config.session_to_visualize})',
                fontsize=18,
                pad=15
            )
            ax.tick_params(
                axis='both',
                which='both',
                labelsize=9,
                rotation=45,
                labelrotation=45
            )
            
            # Dynamic layout adjustments
            plt.tight_layout(pad=2.5)
            
            return DataVisualizer._configure_figure_for_pdf(fig)

        except Exception as e:
            print(f"❌ Heatmap generation failed: {str(e)}")
            return None

    @staticmethod
    def show_statistics(df: pd.DataFrame, config: VisualizationConfig) -> plt.Figure:
        if df is None or df.empty:
            print("⚠️ No data available for statistics")
            return None

        try:
            numeric_cols = df.select_dtypes(include=np.number).columns
            stats = df[numeric_cols].describe().T
            stats['IQR'] = stats['75%'] - stats['25%']
            stats['Outlier_Count'] = 0
            
            for col in numeric_cols:
                outlier_col = f'{col}_Outlier'
                if outlier_col in df.columns:
                    stats.loc[col, 'Outlier_Count'] = df[outlier_col].sum()

            # Get the data we'll display
            display_data = stats[['mean', 'std', 'min', '50%', 'max', 'IQR', 'Outlier_Count']].round(2)
            
            # Create figure
            fig, ax = plt.subplots(figsize=(12, 8))
            ax.axis('off')
            
            # Calculate exact dimensions
            n_rows, n_cols = display_data.shape
            
            # Generate cell colors
            cell_colors = []
            for i in range(n_rows):
                for j in range(n_cols):
                    if j < 2:  # Mean and Std columns
                        norm_val = (display_data.iloc[i,j] - display_data.iloc[:,:2].values.min()) / \
                                (display_data.iloc[:,:2].values.max() - display_data.iloc[:,:2].values.min())
                        cell_colors.append(plt.cm.Blues(norm_val * 0.7 + 0.3))
                    else:  # Other columns
                        norm_val = (display_data.iloc[i,j] - display_data.iloc[:,2:].values.min()) / \
                                (display_data.iloc[:,2:].values.max() - display_data.iloc[:,2:].values.min())
                        cell_colors.append(plt.cm.Reds(norm_val * 0.7 + 0.3))
            
            # Create table
            table = ax.table(
                cellText=display_data.values,
                rowLabels=display_data.index,
                colLabels=display_data.columns,
                cellColours=np.array(cell_colors).reshape(n_rows, n_cols, 4),  # Reshape to match data
                loc='center',
                colWidths=[0.12]*n_cols
            )
            
            # Formatting
            table.auto_set_font_size(False)
            table.set_fontsize(10)
            table.scale(1, 1.5)
            plt.title("Enhanced Statistics", pad=20, fontsize=14)
            plt.tight_layout()
            
            return DataVisualizer._configure_figure_for_pdf(fig)

        except Exception as e:
            print(f"❌ Error generating statistics table: {str(e)}")
            return None
        
class ConditionAnalyzer:
    @staticmethod
    def analyze_session(df: pd.DataFrame, config: VisualizationConfig) -> dict:
        """
        Comprehensive analysis of eye tracking session to determine user condition.
        Returns dictionary with analysis results and suggested interpretations.
        """
        # Initial validation
        if df is None or df.empty:
            return {
                'error': 'No data available for analysis',
                'basic_metrics': {},
                'condition_indicators': {},
                'interpretation': {'primary_conditions': [{
                    'condition': 'Invalid Data',
                    'confidence': 'high',
                    'description': 'No valid eye tracking data was provided'
                }]},
                'recommendations': [{
                    'type': 'data',
                    'priority': 'high',
                    'text': 'Re-run eye tracking session to collect valid data'
                }]
            }
        
        # Check for sufficient movement data
        movement_cols = [c for c in df.columns if 'Magnitude' in c and not c.endswith('_Outlier')]
        if not any(df[col].mean() > 1.0 for col in movement_cols):
            return {
                'error': 'Insufficient eye movement detected',
                'basic_metrics': ConditionAnalyzer._calculate_basic_metrics(df, config),
                'condition_indicators': {},
                'interpretation': {'primary_conditions': [{
                    'condition': 'Minimal Eye Movement',
                    'confidence': 'high',
                    'description': 'Very little eye movement detected during session'
                }]},
                'recommendations': [{
                    'type': 'procedure',
                    'priority': 'high',
                    'text': 'Ensure proper eye tracking calibration and participant engagement'
                }]
            }
        
        # Proceed with normal analysis if data is valid
        analysis = {
            'basic_metrics': {},
            'condition_indicators': {},
            'interpretation': {},
            'recommendations': []
        }
        
        # Calculate basic metrics
        analysis['basic_metrics'] = ConditionAnalyzer._calculate_basic_metrics(df, config)
        
        # Detect specific conditions
        analysis.update(ConditionAnalyzer._detect_conditions(df, config))
        
        # Generate interpretation
        analysis['interpretation'] = ConditionAnalyzer._generate_interpretation(analysis)
        
        # Generate recommendations
        analysis['recommendations'] = ConditionAnalyzer._generate_recommendations(analysis)
        
        return analysis

    @staticmethod
    def _calculate_basic_metrics(df: pd.DataFrame, config: VisualizationConfig) -> dict:
        """Calculate fundamental metrics from the eye tracking data"""
        metrics = {}
        
        # Time metrics
        metrics['duration_seconds'] = df[config.x_axis_column].max() - df[config.x_axis_column].min()
        
        # Movement metrics (using magnitude columns)
        movement_cols = [c for c in df.columns if 'Magnitude' in c and not c.endswith('_Outlier')]
        for col in movement_cols:
            metrics[f'{col}_mean'] = df[col].mean()
            metrics[f'{col}_max'] = df[col].max()
            metrics[f'{col}_std'] = df[col].std()
        
        # Frequency metrics
        freq_cols = [c for c in df.columns if 'Freq_Hz' in c]
        for col in freq_cols:
            metrics[f'{col}_mean'] = df[col].mean()
            metrics[f'{col}_max'] = df[col].max()
        
        # Variance metrics (jitter/tremor)
        var_cols = [c for c in df.columns if 'Variance' in c]
        for col in var_cols:
            metrics[f'{col}_mean'] = df[col].mean()
        
        return metrics

    @staticmethod
    def _detect_conditions(df: pd.DataFrame, config: VisualizationConfig) -> dict:
        """Detect specific conditions based on eye movement patterns"""
        conditions = {
            'condition_indicators': {},
            'flags': []
        }
        
        # 1. Fatigue detection
        left_fatigue, right_fatigue = ConditionAnalyzer._detect_fatigue(df)
        conditions['condition_indicators']['fatigue_left'] = left_fatigue
        conditions['condition_indicators']['fatigue_right'] = right_fatigue
        if left_fatigue['is_present'] or right_fatigue['is_present']:
            conditions['flags'].append('fatigue')
        
        # 2. Nystagmus detection
        nystagmus = ConditionAnalyzer._detect_nystagmus(df)
        conditions['condition_indicators']['nystagmus'] = nystagmus
        if nystagmus['is_present']:
            conditions['flags'].append('nystagmus')
        
        # 3. Convergence insufficiency
        convergence = ConditionAnalyzer._detect_convergence_issues(df)
        conditions['condition_indicators']['convergence_insufficiency'] = convergence
        if convergence['is_present']:
            conditions['flags'].append('convergence_insufficiency')
        
        # 4. Asymmetry detection
        asymmetry = ConditionAnalyzer._detect_asymmetry(df)
        conditions['condition_indicators']['asymmetry'] = asymmetry
        if asymmetry['is_present']:
            conditions['flags'].append('asymmetry')
        
        return conditions

    @staticmethod
    def _detect_fatigue(df: pd.DataFrame) -> tuple:
        """Detect signs of eye fatigue with robust division handling"""
        def safe_divide(a, b, default=1.0):
            """Handle division by zero cases"""
            try:
                return a / b if abs(b) > 1e-6 else default
            except:
                return default
        
        # Calculate left eye metrics with zero-division protection
        left_initial_vel = df['Left_Vel_Magnitude_mm_s'].iloc[:10].mean()
        left_final_vel = df['Left_Vel_Magnitude_mm_s'].iloc[-10:].mean()
        left_initial_var = df['Left_DX_Variance'].iloc[:10].mean()
        left_final_var = df['Left_DX_Variance'].iloc[-10:].mean()
        
        left_metrics = {
            'blink_rate': df['Left_Freq_Hz'].mean(),
            'velocity_decline': safe_divide(left_final_vel, left_initial_vel),
            'variance_increase': safe_divide(left_final_var, left_initial_var) if left_initial_var > 1e-6 else 1.0
        }
        
        # Calculate right eye metrics with zero-division protection
        right_initial_vel = df['Right_Vel_Magnitude_mm_s'].iloc[:10].mean()
        right_final_vel = df['Right_Vel_Magnitude_mm_s'].iloc[-10:].mean()
        right_initial_var = df['Right_DX_Variance'].iloc[:10].mean()
        right_final_var = df['Right_DX_Variance'].iloc[-10:].mean()
        
        right_metrics = {
            'blink_rate': df['Right_Freq_Hz'].mean(),
            'velocity_decline': safe_divide(right_final_vel, right_initial_vel),
            'variance_increase': safe_divide(right_final_var, right_initial_var) if right_initial_var > 1e-6 else 1.0
        }
        
        # Determine fatigue presence with additional checks
        left_fatigue = {
            'is_present': (left_metrics['velocity_decline'] < 0.7 or 
                        left_metrics['variance_increase'] > 1.5) and
                        left_initial_vel > 1.0,  # Only consider if there was meaningful initial movement
            'metrics': left_metrics
        }
        
        right_fatigue = {
            'is_present': (right_metrics['velocity_decline'] < 0.7 or 
                        right_metrics['variance_increase'] > 1.5) and
                        right_initial_vel > 1.0,  # Only consider if there was meaningful initial movement
            'metrics': right_metrics
        }
        
        return left_fatigue, right_fatigue

    @staticmethod
    def _detect_nystagmus(df: pd.DataFrame) -> dict:
        """Detect involuntary rhythmic eye movements"""
        # Look for high frequency with consistent pattern
        left_freq = df['Left_Freq_Hz'].mean()
        right_freq = df['Right_Freq_Hz'].mean()
        
        # Check for rhythmic patterns in velocity
        left_vel_autocorr = df['Left_VX_Norm_Smoothed'].autocorr(lag=5)
        right_vel_autocorr = df['Right_VX_Norm_Smoothed'].autocorr(lag=5)
        
        return {
            'is_present': ((left_freq > 3.0 and abs(left_vel_autocorr) > 0.5) or 
                          (right_freq > 3.0 and abs(right_vel_autocorr) > 0.5)),
            'left_frequency_hz': left_freq,
            'right_frequency_hz': right_freq,
            'left_autocorrelation': left_vel_autocorr,
            'right_autocorrelation': right_vel_autocorr
        }

    @staticmethod
    def _detect_convergence_issues(df: pd.DataFrame) -> dict:
        """Detect problems with eye coordination"""
        # Compare left and right eye movements
        divergence = (df['Left_DX_Norm'] - df['Right_DX_Norm']).abs().mean()
        
        return {
            'is_present': divergence > 0.15,  # Threshold for significant divergence
            'mean_divergence': divergence,
            'max_divergence': (df['Left_DX_Norm'] - df['Right_DX_Norm']).abs().max()
        }

    @staticmethod
    def _detect_asymmetry(df: pd.DataFrame) -> dict:
        """Detect asymmetry between left and right eye movements"""
        vel_diff = (df['Left_Vel_Magnitude_mm_s'] - df['Right_Vel_Magnitude_mm_s']).abs().mean()
        freq_diff = (df['Left_Freq_Hz'] - df['Right_Freq_Hz']).abs().mean()
        
        return {
            'is_present': vel_diff > 5.0 or freq_diff > 0.5,  # mm/s or Hz thresholds
            'velocity_difference_mm_s': vel_diff,
            'frequency_difference_hz': freq_diff
        }

    @staticmethod
    def _generate_interpretation(analysis: dict) -> dict:
        """Generate human-readable interpretation of the analysis"""
        interpretation = {
            'primary_conditions': [],
            'secondary_findings': []
        }
        
        flags = analysis.get('flags', [])
        
        if 'fatigue' in flags:
            interpretation['primary_conditions'].append({
                'condition': 'Eye Fatigue',
                'confidence': 'high',
                'description': 'Signs of eye muscle fatigue detected, showing reduced velocity and increased jitter over time.'
            })
        
        if 'nystagmus' in flags:
            interpretation['primary_conditions'].append({
                'condition': 'Nystagmus',
                'confidence': 'medium',
                'description': 'Rhythmic, involuntary eye movements detected, which may indicate vestibular or neurological issues.'
            })
        
        if 'convergence_insufficiency' in flags:
            interpretation['primary_conditions'].append({
                'condition': 'Convergence Insufficiency',
                'confidence': analysis['condition_indicators']['convergence_insufficiency']['mean_divergence'] > 0.2 and 'high' or 'medium',
                'description': 'Eyes show difficulty maintaining proper alignment during tracking tasks.'
            })
        
        if 'asymmetry' in flags:
            interpretation['primary_conditions'].append({
                'condition': 'Eye Movement Asymmetry',
                'confidence': 'medium',
                'description': 'Significant differences detected between left and right eye movement patterns.'
            })
        
        if not interpretation['primary_conditions']:
            interpretation['primary_conditions'].append({
                'condition': 'Normal Eye Movement',
                'confidence': 'high',
                'description': 'No significant abnormalities detected in eye movement patterns.'
            })
        
        return interpretation

    @staticmethod
    def _generate_recommendations(analysis: dict) -> list:
        """Generate recommendations based on the analysis"""
        recommendations = []
        flags = analysis.get('flags', [])
        
        if 'fatigue' in flags:
            recommendations.append({
                'type': 'lifestyle',
                'priority': 'high',
                'text': 'Take regular breaks from screen time (20-20-20 rule: every 20 minutes, look at something 20 feet away for 20 seconds)'
            })
            recommendations.append({
                'type': 'medical',
                'priority': 'medium',
                'text': 'Consider consulting an optometrist if fatigue symptoms persist'
            })
        
        if 'nystagmus' in flags:
            recommendations.append({
                'type': 'medical',
                'priority': 'high',
                'text': 'Consult with a neurologist or ophthalmologist for evaluation of nystagmus'
            })
        
        if 'convergence_insufficiency' in flags:
            recommendations.append({
                'type': 'therapy',
                'priority': 'medium',
                'text': 'Consider vision therapy exercises to improve eye coordination'
            })
        
        if 'asymmetry' in flags:
            recommendations.append({
                'type': 'medical',
                'priority': 'medium',
                'text': 'Further evaluation recommended to determine cause of eye movement asymmetry'
            })
        
        if not recommendations:
            recommendations.append({
                'type': 'maintenance',
                'priority': 'low',
                'text': 'Continue current eye care routine, no specific recommendations needed'
            })
        
        return recommendations



# --- PDF Report Generator ---
class ReportGenerator:
    @staticmethod
    def _add_figure_to_story(fig: plt.Figure, title: str, styles, temp_dir: str) -> tuple:
        """Helper method to save figure and create report elements"""
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, f"{title.replace(' ', '_')}.png")
        
        try:
            # Save with higher DPI and tight layout
            fig.savefig(temp_path, dpi=300, bbox_inches='tight', facecolor='white')
            plt.close(fig)
            
            if os.path.exists(temp_path):
                # Create PDF elements
                elements = [
                    Paragraph(title, styles['Heading2']),
                    Spacer(1, 0.1*inch),
                    Image(temp_path, width=6*inch, height=4*inch, kind='proportional'),
                    Spacer(1, 0.2*inch)
                ]
                return elements, temp_path
        
        except Exception as e:
            print(f"⚠️ Error saving figure {title}: {str(e)}")
            return [Paragraph(f"⚠️ Could not generate {title}", styles['Normal'])], None
        
        return [], None

    @staticmethod
    def _add_visualizations(df: pd.DataFrame, config: VisualizationConfig, styles) -> list:
        """Improved visualization handling for PDF"""
        elements = []
        temp_files = []
        temp_dir = r"Progress-Reports-Generation\temp_report_images"
        
        visualization_sequence = [
            ('Statistical Summary', DataVisualizer.show_statistics),
            ('Feature Trends', DataVisualizer.create_grid_plot),
            ('Feature Correlations', DataVisualizer.create_correlation_heatmap),
            ('Movement Analysis', DataVisualizer.plot_movement_analysis)
        ]
        
        for title, plot_method in visualization_sequence:
            try:
                fig = plot_method(df, config)
                if fig:
                    fig_elements, temp_path = ReportGenerator._add_figure_to_story(
                        fig, title, styles, temp_dir
                    )
                    elements.extend(fig_elements)
                    if temp_path:
                        temp_files.append(temp_path)
            except Exception as e:
                elements.append(Paragraph(
                    f"⚠️ Error generating {title}: {str(e)}",
                    styles['Normal']
                ))
        
        # Add automatic cleanup (optional)
        # for file in temp_files:
        #     try:
        #         os.remove(file)
        #     except:
        #         pass
                
        return elements

    
    @staticmethod
    def generate_session_report(df: pd.DataFrame, config: VisualizationConfig, output_path: str):
        """Generate one PDF report for a single session (sheet) within an excel file.
        PDF contains relevant eye tracking parameter visualizations as well as condition analysis"""
        # Set matplotlib backend and style
        plt.switch_backend('Agg')  # Non-interactive backend
        plt.style.use(config.plot_style)
        
        # Perform condition analysis
        condition_analysis = ConditionAnalyzer.analyze_session(df, config)
        
        # Create document template
        doc = SimpleDocTemplate(output_path, pagesize=A4,
                              rightMargin=72, leftMargin=72,
                              topMargin=72, bottomMargin=72)
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1  # Center
        )
        
        # Story holds all elements to be added to PDF
        story = []
        
        # 1. Cover Page
        story.append(Paragraph("Eye Tracking Analysis Report", title_style))
        story.append(Spacer(1, 0.5*inch))
        story.append(Paragraph(f"Session: {config.session_to_visualize}", styles['Heading2']))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 
                             styles['Normal']))
        story.append(PageBreak())
        
        print("111")
        
        # 2. Condition Summary
        story += ReportGenerator._create_condition_summary(condition_analysis, styles)
        story.append(PageBreak())
        
        print("222")

        # 3. Detailed Analysis
        story += ReportGenerator._create_detailed_analysis(condition_analysis, styles)
        story.append(PageBreak())
        
        print("333")
        
        # 4. Visualizations
        story += ReportGenerator._add_visualizations(df, config, styles)
        
        # Build the PDF
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )
        doc.build(story)
                    
    
    @staticmethod
    def _create_condition_summary(analysis: dict, styles) -> list:
        """Create condition summary section"""
        elements = []
        
        elements.append(Paragraph("Condition Summary", styles['Heading1']))
        elements.append(Spacer(1, 0.25*inch))
        
        if 'error' in analysis:
            elements.append(Paragraph(analysis['error'], styles['Heading2']))
            return elements
        
        # Primary Findings
        elements.append(Paragraph("Primary Findings:", styles['Heading2']))
        for condition in analysis['interpretation']['primary_conditions']:
            elements.append(Paragraph(
                f"• {condition['condition']} ({condition['confidence']} confidence)",
                styles['Bullet']
            ))
            elements.append(Paragraph(
                f"  {condition['description']}",
                styles['Normal']
            ))
            elements.append(Spacer(1, 0.1*inch))
        
        # Recommendations
        elements.append(Paragraph("Recommendations:", styles['Heading2']))
        for rec in analysis['recommendations']:
            text_color = colors.red if rec['priority'] == 'high' else colors.black
            style = ParagraphStyle(
                'RecStyle',
                parent=styles['Normal'],
                textColor=text_color,
                leftIndent=20
            )
            elements.append(Paragraph(
                f"• [{rec['priority'].upper()}] {rec['text']}",
                style
            ))
            elements.append(Spacer(1, 0.05*inch))
        
        return elements
    
    @staticmethod
    def _create_detailed_analysis(analysis: dict, styles) -> list:
        """Create detailed metrics page for the condition analysis"""
        elements = []
        
        elements.append(Paragraph("Detailed Analysis", styles['Heading1']))
        elements.append(Spacer(1, 0.25*inch))
        
        if 'error' in analysis:
            return elements
        
        # Create table data - convert all text to Paragraphs for proper wrapping
        table_data = [
            [
                Paragraph("Metric", styles['Normal']),
                Paragraph("Left Eye", styles['Normal']),
                Paragraph("Right Eye", styles['Normal']),
                Paragraph("Threshold", styles['Normal']),
                Paragraph("Interpretation", styles['Normal'])
            ]
        ]
        
        # Helper function to create wrapped cell content
        def create_cell(content, style=styles['Normal']):
            return Paragraph(content, style) if isinstance(content, str) else content
        
        # Add movement metrics with wrapped text
        table_data.append([
            create_cell("Movement Velocity (mm/s)", styles['Normal']),
            create_cell(f"{analysis['basic_metrics'].get('Left_Vel_Magnitude_mm_s_mean', 0):.1f}"),
            create_cell(f"{analysis['basic_metrics'].get('Right_Vel_Magnitude_mm_s_mean', 0):.1f}"),
            create_cell(">15 normal"),
            create_cell("Low velocity may indicate fatigue or neurological impairment")
        ])
        
        # Add frequency metrics
        table_data.append([
            create_cell("Movement Frequency (Hz)"),
            create_cell(f"{analysis['basic_metrics'].get('Left_Freq_Hz_mean', 0):.1f}"),
            create_cell(f"{analysis['basic_metrics'].get('Right_Freq_Hz_mean', 0):.1f}"),
            create_cell("0.5-2.5 normal"),
            create_cell("High frequency may indicate nystagmus or other oscillatory conditions")
        ])
        
        # Add variance metrics
        table_data.append([
            create_cell("Position Variance"),
            create_cell(f"{analysis['basic_metrics'].get('Left_DX_Variance_mean', 0):.3f}"),
            create_cell(f"{analysis['basic_metrics'].get('Right_DX_Variance_mean', 0):.3f}"),
            create_cell("<0.01 stable"),
            create_cell("High variance indicates tremor, instability, or poor tracking ability")
        ])
        
        # Column widths (adjust as needed)
        col_widths = [1.5*inch, 0.8*inch, 0.8*inch, 1.0*inch, 2.0*inch]
        
        # Create table with dynamic row heights
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            
            # Data row styling
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            
            # Content handling for dynamic heights
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('WORDWRAP', (0,0), (-1,-1), True),
            
            # Numeric alignment
            ('ALIGN', (1,1), (2,-1), 'CENTER'),
            ('ALIGN', (3,1), (3,-1), 'CENTER'),
            
            # Padding for better readability
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Fix the condition flags section:
        if analysis.get('flags'):
            elements.append(Paragraph("Condition Flags Detected:", styles['Heading2']))
            
            # Handle case where flags is a list of strings
            if isinstance(analysis['flags'], list):
                for flag in analysis['flags']:
                    elements.append(Paragraph(
                        f"• {flag.replace('_', ' ').title()}",
                        styles['Bullet']
                    ))
            # Handle case where flags is a list of dicts (if needed)
            elif isinstance(analysis['flags'][0], dict):
                for flag in analysis['flags']:
                    elements.append(Paragraph(
                        f"• {flag.get('condition', 'Unknown')}: {flag.get('description', '')}",
                        styles['Bullet']
                    ))
            
            elements.append(Spacer(1, 0.2*inch))
        
        return elements

# --- Main Execution ---
def generate_all_reports(input_file_path: str, output_dir: str = "reports"):
    """Generate PDF reports for all or selected sessions in an Excel file."""
    os.makedirs(output_dir, exist_ok=True)
    
    if not input_file_path.endswith('.xlsx'):
        raise ValueError("Only Excel files (.xlsx) with multiple sheets are supported")
    
    wb = load_workbook(input_file_path, read_only=True)
    session_sheets = [sheet for sheet in wb.sheetnames if sheet.startswith('Session_')]
    
    if not session_sheets:
        print("⚠️ No session sheets found (sheets should start with 'Session_')")
        return
    
    print(f"\n📊 Found {len(session_sheets)} sessions in the file:")
    for i, sheet in enumerate(session_sheets, 1):
        print(f"{i}. {sheet}")
    
    while True:
        selection = input("\nChoose sessions to process (e.g., '1-3,5', 'all', or specific numbers): ").strip().lower()
        
        if selection == 'all':
            selected_sheets = session_sheets
            break
        else:
            try:
                selected_indices = set()
                parts = selection.split(',')
                for part in parts:
                    if '-' in part:
                        start, end = map(int, part.split('-'))
                        selected_indices.update(range(start, end + 1))
                    else:
                        selected_indices.add(int(part))
                
                # Validate indices
                valid_indices = range(1, len(session_sheets) + 1)
                if all(idx in valid_indices for idx in selected_indices):
                    selected_sheets = [session_sheets[i-1] for i in sorted(selected_indices)]
                    break
                else:
                    print("⚠️ Some numbers are out of range. Please try again.")
            except ValueError:
                print("⚠️ Invalid input. Please enter numbers separated by commas or ranges with hyphens.")
    
    print(f"\n🔍 Processing {len(selected_sheets)} selected session(s)...")
    for sheet_name in selected_sheets:
        print(f"\nProcessing {sheet_name}...")
        try:
            config = VisualizationConfig(
                excel_file_path=input_file_path,
                session_to_visualize=sheet_name
            )
            df = DataLoader.load_session_data(input_file_path, sheet_name, config)   # Load Dataframe
            if df is not None:  # Only generate report if data loaded successfully
                ReportGenerator.generate_session_report(
                    df,  
                    config,
                    os.path.join(output_dir, f"Report_{sheet_name}.pdf")
                )
                print(f"✅ Successfully generated report for {sheet_name}")
            else:
                print(f"❌ Skipping report generation for {sheet_name} - no valid data")
        except Exception as e:
            print(f"❌ Error processing {sheet_name}: {str(e)}")
    
    print("\n🎉 Report generation completed!")

# --- Usage Example ---
if __name__ == "__main__":
    generate_all_reports(
        input_file_path=r".\LiveData.xlsx",
        output_dir=r"Progress-Reports-Generation\session_reports"
    )