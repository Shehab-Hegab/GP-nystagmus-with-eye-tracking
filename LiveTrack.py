
import cv2
import numpy as np
import datetime
import mediapipe as mp
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from joblib import dump, load
import tkinter as tk
from tkinter import filedialog

# Suppress TensorFlow warnings
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

# Initialize MediaPipe Face Mesh
mp_face_mesh = mp.solutions.face_mesh
face_mesh = mp_face_mesh.FaceMesh(
    static_image_mode=False,
    max_num_faces=1,
    refine_landmarks=True,
    min_detection_confidence=0.2,
    min_tracking_confidence=0.2,
)

# -----------------------------------------------
# OPTION 1 (commented out by default): Use webcam
cap = cv2.VideoCapture(0)
if not cap.isOpened():
    print("Error: Could not open webcam.")
    exit()

# OPTION 2: Use a file dialog to select any video
root = tk.Tk()

root.withdraw()  # Hide the main tkinter window

video_path = filedialog.askopenfilename(
    title="Select a video file",
    filetypes=[("Video Files", "*.mp4 *.avi *.mov *.mkv"), ("All Files", "*.*")]
)

if not video_path:
    print("No file selected. Exiting...")
    exit()

cap = cv2.VideoCapture(video_path)
if not cap.isOpened():
    print("Error: Could not open video file.")
    exit()
# -----------------------------------------------

# Function to calculate Euclidean distance between two points
def euclidean_distance(point1, point2):
    return np.sqrt((point1[0] - point2[0])**2 + (point1[1] - point2[1])**2)

# Initialize measurement variables
previous_displacement_left, previous_displacement_right = 0, 0
previous_distance_left_by_mm, previous_distance_right_by_mm = 0, 0
oscillation_count_left, oscillation_count_right = 0, 0
frame_index = 0
Time = 0
Time_with_face = 0
previous_Time = 0

# Model and data paths
MODEL_PATH = 'tremor_model.joblib'
SCALER_PATH = 'tremor_scaler.joblib'
DATA_PATH = r"F:\GP\ML\LiveData\LiveData.xlsx"

# Initialize model and scaler
model = None
scaler = None

# Try to load a pre-trained model
try:
    if os.path.exists(MODEL_PATH) and os.path.exists(SCALER_PATH):
        model = load(MODEL_PATH)
        scaler = load(SCALER_PATH)
        print("Loaded pre-trained model successfully!")
    else:
        raise FileNotFoundError
except Exception as e:
    print(f"No pre-trained model found: {e}. Training a new model...")
    
    # Check if training data exists
    if os.path.exists(DATA_PATH):
        # Load and prepare data
        df = pd.read_excel(DATA_PATH)
        
        # Add Label column if missing
        if 'Label' not in df.columns:
            print("Adding 'Label' column with default value 0")
            df['Label'] = 0  # Default value (modify as needed)
        
        # Drop rows with missing values
        df = df.dropna()
        
        # Define features
        features = [
            'Left Move (mm)', 'Right Move (mm)',
            'Left Velocity (mm/s)', 'Right Velocity (mm/s)',
            'Left Frequency (Hz)', 'Right Frequency (Hz)'
        ]
        
        # Ensure all features exist in the DataFrame
        for feature in features:
            if feature not in df.columns:
                df[feature] = 0
        
        X = df[features]
        y = df['Label']
        
        # Split the dataset into training and testing sets
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=0.2, random_state=42
        )
        
        # Scale features
        scaler = StandardScaler()
        X_train_scaled = scaler.fit_transform(X_train)
        X_test_scaled = scaler.transform(X_test)
        
        # Train the RandomForest model
        model = RandomForestClassifier(
            n_estimators=100,
            max_depth=5,
            random_state=42
        )
        model.fit(X_train_scaled, y_train)
        
        # Save the trained model and scaler
        dump(model, MODEL_PATH)
        dump(scaler, SCALER_PATH)
        print(f"Model trained with accuracy: {model.score(X_test_scaled, y_test):.2f}")
    else:
        print("No training data available. Real-time predictions are disabled.")

# Data collection structure
data = [["Index", "Time (s)", "TimeWithFace", "LeftMove(mm)", "RightMove(mm)", 
         "LeftVel(mm/s)", "RightVel(mm/s)", "LeftFreq(Hz)", "RightFreq(Hz)", 
         "LeftAmp(mm)", "RightAmp(mm)", "Prediction"]]

# Main processing loop
while True:
    ret, frame = cap.read()
    if not ret:
        print("Error: Failed to read frame (webcam or video).")
        break
    
    # Flip the frame horizontally (optional)
    frame = cv2.flip(frame, 1)
    
    # Convert to RGB color space for MediaPipe
    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    # Process the frame to get face landmarks
    results = face_mesh.process(rgb_frame)
    
    # Calculate time per frame (frame_time).
    fps = cap.get(cv2.CAP_PROP_FPS) or 30
    frame_time = 1 / fps
    
    prediction_text = "No face detected"
    current_features = []
    
    # -------------------------------------------------------------------
    # Initialize these to 0 before checking if a face is found
    displacement_left = 0
    displacement_right = 0
    velocity_left = 0
    velocity_right = 0
    amplitude_left = 0
    amplitude_right = 0
    frequency_left = 0
    frequency_right = 0
    # -------------------------------------------------------------------
    
    # If a face is detected
    if results.multi_face_landmarks:
        for face_landmarks in results.multi_face_landmarks:
            h, w, _ = frame.shape
            
            # Nose landmark (index 6 in MediaPipe Face Mesh)
            top_nose = (
                int(face_landmarks.landmark[6].x * w),
                int(face_landmarks.landmark[6].y * h)
            )
            
            # Left eye landmarks
            left_eye_center = (
                int(face_landmarks.landmark[468].x * w),
                int(face_landmarks.landmark[468].y * h)
            )
            left_eye_left = (
                int(face_landmarks.landmark[471].x * w),
                int(face_landmarks.landmark[471].y * h)
            )
            left_eye_right = (
                int(face_landmarks.landmark[469].x * w),
                int(face_landmarks.landmark[469].y * h)
            )
            
            # Calculate left eye metrics
            left_radius = euclidean_distance(left_eye_left, left_eye_right)
            # Prevent division by zero if left_radius is extremely small
            if left_radius < 1e-6:
                left_radius = 1e-6
            
            left_dist = euclidean_distance(top_nose, left_eye_center)
            left_mm = left_dist * 12 / left_radius  # or keep in pixels
            
            # Right eye landmarks
            right_eye_center = (
                int(face_landmarks.landmark[473].x * w),
                int(face_landmarks.landmark[473].y * h)
            )
            right_eye_left = (
                int(face_landmarks.landmark[476].x * w),
                int(face_landmarks.landmark[476].y * h)
            )
            right_eye_right = (
                int(face_landmarks.landmark[474].x * w),
                int(face_landmarks.landmark[474].y * h)
            )
            
            # Calculate right eye metrics
            right_radius = euclidean_distance(right_eye_left, right_eye_right)
            if right_radius < 1e-6:
                right_radius = 1e-6
            
            right_dist = euclidean_distance(top_nose, right_eye_center)
            right_mm = right_dist * 12 / right_radius
            
            # Calculate dynamics
            displacement_left = left_mm - previous_distance_left_by_mm
            displacement_right = right_mm - previous_distance_right_by_mm
            
            velocity_left = abs(displacement_left / frame_time) if frame_time else 0
            velocity_right = abs(displacement_right / frame_time) if frame_time else 0
            
            amplitude_left = abs(displacement_left)
            amplitude_right = abs(displacement_right)
            
            # Update oscillation counts (sign changes in displacement)
            if displacement_left * previous_displacement_left < 0:
                oscillation_count_left += 1
            if displacement_right * previous_displacement_right < 0:
                oscillation_count_right += 1
                
            frequency_left = oscillation_count_left / Time_with_face if Time_with_face else 0
            frequency_right = oscillation_count_right / Time_with_face if Time_with_face else 0
            
            # Prepare features for the model
            current_features = [
                amplitude_left, amplitude_right,
                velocity_left, velocity_right,
                frequency_left, frequency_right
            ]
            
            # Make a prediction if the model and scaler are loaded
            if model and scaler and len(current_features) == 6:
                try:
                    scaled_features = scaler.transform([current_features])
                    pred = model.predict(scaled_features)
                    proba = model.predict_proba(scaled_features)[0]
                    prediction_text = f"State: {pred[0]} ({np.max(proba):.2f})"
                except Exception as e:
                    prediction_text = f"Prediction error: {str(e)}"
            
            # Update previous values for the next frame
            previous_distance_left_by_mm = left_mm
            previous_distance_right_by_mm = right_mm
            previous_displacement_left = displacement_left
            previous_displacement_right = displacement_right
            Time_with_face += frame_time
    
    # Collect data for each frame
    data.append([
        frame_index,
        Time,
        Time_with_face,
        displacement_left,
        displacement_right,
        velocity_left,
        velocity_right,
        frequency_left,
        frequency_right,
        amplitude_left,
        amplitude_right,
        prediction_text
    ])
    
    # Display the prediction on the frame
    cv2.putText(frame, prediction_text, (10, 30),
                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
    
    # Show the video feed
    cv2.imshow("Tremor Detection", frame)
    
    # Update time and frame index
    Time += frame_time
    frame_index += 1
    
    # Exit if 'Esc' is pressed
    if cv2.waitKey(1) & 0xFF == 27:
        break

# Save results to Excel
output_dir = r"F:\GP\ML\LiveData"
os.makedirs(output_dir, exist_ok=True)
output_file = os.path.join(output_dir, "LiveData.xlsx")

try:
    if not os.path.exists(output_file):
        Workbook().save(output_file)
        
    wb = load_workbook(output_file)
    sheet_name = f"Session_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
    ws = wb.create_sheet(title=sheet_name)
    
    for row in data:
        ws.append(row)
        
    wb.save(output_file)
    print(f"Data saved successfully to {sheet_name}")
    
except Exception as e:
    print(f"Error saving data: {str(e)}")

# Cleanup resources
cap.release()
cv2.destroyAllWindows()